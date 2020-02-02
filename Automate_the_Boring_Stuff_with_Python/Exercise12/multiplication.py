# Python3
# multiplication.py - NxN乘法表

import os
import openpyxl
from openpyxl.styles import Font, Style

os.chdir(os.path.join(os.getcwd(), 'Exercise12/'))

fontObj1 = Font(name='Times New Roman', bold=True)
styleObj1 = Style(fontObj1)



wb = openpyxl.Workbook()
wb.create_sheet(index=0, title='mutiplication')
# print(wb.get_sheet_names())
sheet = wb.get_sheet_by_name('mutiplication')

print('The number of NxN :')
N = input()
# Creat data in column A and row 1
for i in range(2, N+2):
    sheet['A' + str(i)] = i-1
    sheet['A' + str(i)].style = styleObj1
    for j in range(2, N+2):
        sheet.cell(row=1, column=j).value = j-1
        sheet.cell(row=1, column=j).style = styleObj1
        # multiple
        sheet.cell(row=i, column=j).value = sheet['A'+str(i)].value * sheet.cell(row=1, column=j).value



# sheet.cell(row=i, column=j).value = sheet['A'+str(i)] * sheet.cell(row=1, column=j)

wb.save('multiplication.xlsx')