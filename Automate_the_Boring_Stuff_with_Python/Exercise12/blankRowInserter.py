# Python3
# blankRowInsert.py - 

import os 
import openpyxl
print(os.getcwd())
os.chdir(os.path.join(os.getcwd(), 'Exercise12/'))

OriFile = '/home/linux/Documents/Python_test/Automate_the_Boring_Stuff_with_Python/12_Project/produceSales.xlsx'
N = 2 # 前N行不变
M = 3 # 插入M个空行
def blaninsert(OriFile, M, N):
    #
    wb_O = openpyxl.load_workbook(OriFile)
    sheet_O = wb_O.get_sheet_by_name('Sheet')

    wb_new = openpyxl.Workbook()
    wb_new.create_sheet(index=0, title='new')
    sheet_new = wb_new.get_sheet_by_name('new')

    for i in range(1, sheet_O.max_row + 1):
        for j in range(1, sheet_O.max_column):
            if i > N:
                sheet_new.cell(row=i+M, column=j).value = sheet_O.cell(row=i, column=j).value
            else:
                sheet_new.cell(row=i, column=j).value = sheet_O.cell(row=i, column=j).value
        
        

    wb_new.save('blanRowInset.xlsx')
    print('Done!')

blaninsert(OriFile, M, N)