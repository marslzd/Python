# 处理Excel电子表格

# Excel文档
# pip3 install openpyxl==2.1.4

import openpyxl

# 读取Excel文档
# 用openpyxl打开Excel文档
wb = openpyxl.load_workbook('example.xlsx')
print(wb.get_sheet_names())

sheet = wb.get_sheet_by_name('Sheet3')
print(sheet)
print(type(sheet))

print(sheet.title)

anotherSheet = wb.get_active_sheet()
print(anotherSheet)

# 从表中取得单元格
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
print(sheet['A1'])

print(sheet['A1'].value)

c = sheet['B1']
print(c.value)
print('Row ' + str(c.row) + ', Column ' + c.column + ' is ' + c.value)
print('Cell ' + c.coordinate + ' is ' + c.value)
print(sheet['C1'].value)

print(sheet.cell(row=1, column=2))
print(sheet.cell(row=1, column=2).value)

for i in range(1, 8):
    print(i, sheet.cell(row=i, column=2).value)

# 打印奇数列
for i in range(1, 8, 2):
    print(i, sheet.cell(row=i, column=2).value)

# 返回表的大小
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
print(sheet.get_highest_row())      # 行
print(sheet.get_highest_column())   # 列

# 列字母和数字之间的转换
# 从字母转换成数字 column_index_from_string
# 从数字转换成字母 get_column_letter
import openpyxl
from openpyxl.cell import get_column_letter, column_index_from_string

print(get_column_letter(1))
print(get_column_letter(2))
print(get_column_letter(27))
print(get_column_letter(900))

wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
print(get_column_letter(sheet.get_highest_column()))

print(column_index_from_string('A'))
print(column_index_from_string('AA'))

# 从表中取得列和行
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
print(tuple(sheet['A1':'C3']))

for rowOfCellObjects in sheet['A1' : 'C3']:
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate, cellObj.value)
    print('--- End OF Row ---')

wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.get_active_sheet()
print(sheet.columns[1])

for cellObj in sheet.columns[1]:
    print(cellObj.value)


# 写入Excel文档
# 创建并保存
wb = openpyxl.Workbook()
print(wb.get_sheet_names())

sheet = wb.get_sheet_by_name('Sheet')
sheet.title = 'Spam Bacon Eggs Sheet'
print(wb.get_sheet_names())

wb = openpyxl.load_workbook('example.xlsx')
print(wb.get_sheet_names())
sheet = wb.get_sheet_by_name('Sheet1')
sheet.title = 'Spam Spam Spam'
wb.save('example_copy.xlsx')

# 创建和删除
wb = openpyxl.Workbook()
print(wb.get_sheet_names())

wb.create_sheet()
print(wb.get_sheet_names())

wb.create_sheet(index=0, title='First Sheet')
print(wb.get_sheet_names())

wb.create_sheet(index=2, title='Middle Sheet')
print(wb.get_sheet_names())

wb.remove_sheet(wb.get_sheet_by_name('Sheet'))
print(wb.get_sheet_names())

sheet = wb.get_sheet_by_name('First Sheet')
sheet['A1'] = 'Hello World'
print(sheet['A1'].value)
# wb.save('/home/linux/Desktop/test.xlsx')

##################################################
# 设置单元格的字体风格
from openpyxl.styles import Font, Style
import openpyxl

wb = openpyxl.Workbook()
sheet = wb.get_sheet_by_name('Sheet')
italic24Font = Font(size=24, italic=True)
styleObj = Style(font=italic24Font)
sheet['A1'].style = styleObj
sheet['A1'] = 'Hello World'

# Font对象
fontObj1 = Font(name='Times New Roman', bold=True, size=24)
styleObj1 = Style(font=fontObj1)
sheet['A2'].style = styleObj1
sheet['A2'] = 'Bold Times New Roman'

fontObj2 = Font(size=24, italic=True)
styleObj2 = Style(fontObj2)
sheet['A3'].style = styleObj2
sheet['A3'] = '24pt Italic'
# wb.save('styled.xlsx')
# 公式
wb = openpyxl.Workbook()
sheet = wb.get_sheet_by_name('Sheet')
sheet['A1'] = 200
sheet['A2'] = 300
sheet['A3'] = '=SUM(A1:A2)'
wb.save('writeFormular.xlsx')

wbFormular = openpyxl.load_workbook('writeFormular.xlsx')
sheet = wbFormular.get_sheet_by_name('Sheet')
print(sheet['A3'].value)

wbDataOnly = openpyxl.load_workbook('writeFormular.xlsx', data_only=True)
sheet = wbDataOnly.get_sheet_by_name('Sheet')
print(sheet['A3'].value)  # 500 

# 调整行和列
wb = openpyxl.Workbook()
sheet = wb.get_sheet_by_name('Sheet')
sheet['A1'] = 'Tall row'
sheet['A2'] = 'Wide column'
sheet.row_dimensions[1].height = 70
sheet.column_dimensions['B'].width = 20

# 合并和拆分单元格
sheet.merge_cells('A2:D4')
sheet['A2'] = 'Twelve cells merged together'
sheet.merge_cells('C5:D5')
sheet['C5'] = 'Two merged cells.'

# sheet.unmerge_cells('A2:D4')

# 冻结窗格
# sheet.freeze_panes = 'A2'  # 冻结行1
# sheet.freeze_panes = 'B1'  # 冻结列A
# sheet.freeze_panes = 'C!'  # 冻结列A和列B
# sheet.freeze_panes = 'C2'  # 冻结行1和列A和列B

# 图表
# import createChart.py
