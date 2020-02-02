# Python3
# updateProduce.py - Corrects cost in produceScales.xlsx

import os
import openpyxl

os.chdir(os.path.join(os.getcwd(), '12_Project/'))

# Read Excel
wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.get_sheet_by_name('Sheet')

# The price types and their updated prices
PRICE_UPDATES = {'Garlic' : 3.07,
                 'Celery' : 1.19,
                 'Lemon'  : 1.27}

# Loop through the rows and update the prices
for rowNum in range(2, sheet.max_row+1): # skip the first row
    produceName = sheet.cell(row=rowNum, column=1).value
    if produceName in PRICE_UPDATES:
        sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]

print('Update Done!')

wb.save('updateproduceSales.xlsx')
print('ALL Done!')